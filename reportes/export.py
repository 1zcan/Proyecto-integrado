"""
export.py (Final Version - Exports all consolidated REM data)

Este módulo maneja la generación de archivos físicos (Excel, PDF)
basado en los datos procesados por services.py.
"""

import io
from django.http import HttpResponse

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch

# ----------------------------------------------------------------------
# EXPORTAR A EXCEL (XLSX)
# ----------------------------------------------------------------------
def export_rem_excel(datos):
    """
    Genera un archivo Excel (XLSX) con todos los datos consolidados del REM.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"REM {datos['periodo']}"
    
    # Estilo base
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3DD", fill_type="solid")
    
    # Título
    ws['B2'] = f"Reporte REM Consolidado - Periodo {datos['periodo']}"
    ws['B2'].font = Font(bold=True, size=14)
    
    current_row = 5
    
    # === 🟢 REM A11: TAMIZAJES MADRE ===
    ws[f'B{current_row}'] = "REM A11 - Tamizajes Madre"
    ws[f'B{current_row}'].font = Font(bold=True)
    current_row += 1
    
    ws[f'B{current_row}'] = "Tamizajes VIH Positivos"
    ws[f'C{current_row}'] = datos['rem_a11']['tamizajes_vih_positivos']
    current_row += 1
    ws[f'B{current_row}'] = "Tamizajes VDRL Positivos"
    ws[f'C{current_row}'] = datos['rem_a11']['tamizajes_vdrl_positivos']
    current_row += 1
    ws[f'B{current_row}'] = "VDRL Positivos con Tratamiento"
    ws[f'C{current_row}'] = datos['rem_a11']['tamizajes_vdrl_con_tratamiento']
    current_row += 1
    ws[f'B{current_row}'] = "Tamizajes Hepatitis B Positivos"
    ws[f'C{current_row}'] = datos['rem_a11']['tamizajes_hepb_positivos']
    current_row += 2 # Espacio
    
    # === 🟢 REM A21: PARTO Y CLASIFICACIÓN ===
    ws[f'B{current_row}'] = "REM A21 - Modelo Atención Parto"
    ws[f'B{current_row}'].font = Font(bold=True)
    current_row += 1
    
    ws[f'B{current_row}'] = "Total Partos Registrados"
    ws[f'C{current_row}'] = datos['rem_a21']['total_partos']
    current_row += 1

    # Desglose dinámico de Tipo de Parto
    ws[f'B{current_row}'] = "Desglose por Tipo de Parto:"
    ws[f'B{current_row}'].font = Font(italic=True)
    current_row += 1
    for desglose in datos['rem_a21']['desglose_partos']:
        cell_b = ws[f'B{current_row}']
        cell_b.value = desglose['tipo_parto__valor']
        cell_b.alignment = Alignment(indent=1)
        ws[f'C{current_row}'] = desglose['total']
        current_row += 1
        
    # Desglose dinámico de Robson
    ws[f'B{current_row}'] = "Desglose por Grupo Robson:"
    ws[f'B{current_row}'].font = Font(italic=True)
    current_row += 1
    for desglose in datos['rem_a21']['desglose_robson']:
        cell_b = ws[f'B{current_row}']
        cell_b.value = f"Grupo {desglose['clasificacion_robson__grupo']}" # Clave corregida
        cell_b.alignment = Alignment(indent=1)
        ws[f'C{current_row}'] = desglose['total']
        current_row += 1
        
    ws[f'B{current_row}'] = "Partos con Acompañamiento"
    ws[f'C{current_row}'] = datos['rem_a21']['partos_con_acompanamiento']
    current_row += 2 # Espacio

    # === 🟢 REM A24: RECIÉN NACIDO Y CALIDAD ===
    ws[f'B{current_row}'] = "REM A24 - Atención Recién Nacido"
    ws[f'B{current_row}'].font = Font(bold=True)
    current_row += 1
    
    ws[f'B{current_row}'] = "Total Recién Nacidos"
    ws[f'C{current_row}'] = datos['rem_a24']['total_rn']
    current_row += 1
    
    ws[f'B{current_row}'] = "RN con Lactancia (60 min)"
    ws[f'C{current_row}'] = datos['rem_a24']['rn_con_lm_60min']
    current_row += 1
    ws[f'B{current_row}'] = "RN con Bajo Peso (<2500g)"
    ws[f'C{current_row}'] = datos['rem_a24']['rn_bajo_peso']
    current_row += 1
    ws[f'B{current_row}'] = "RN con Apgar < 7 (5 min)"
    ws[f'C{current_row}'] = datos['rem_a24']['rn_apgar_bajo']
    current_row += 1
    ws[f'B{current_row}'] = "RN con Reanimación"
    ws[f'C{current_row}'] = datos['rem_a24']['rn_con_reanimacion']
    current_row += 1
    ws[f'B{current_row}'] = "Profilaxis Vitamina K"
    ws[f'C{current_row}'] = datos['rem_a24']['rn_con_vitamina_k']
    current_row += 1
    ws[f'B{current_row}'] = "Profilaxis Oftálmica"
    ws[f'C{current_row}'] = datos['rem_a24']['rn_con_prof_oftalmica']

    # Configurar respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="rem_{datos["periodo"]}.xlsx"'},
    )
    wb.save(response)
    return response


# ----------------------------------------------------------------------
# EXPORTAR A PDF
# ----------------------------------------------------------------------
def export_rem_pdf(datos):
    """
    Genera un archivo PDF con todos los datos consolidados del REM.
    """
    buffer = io.BytesIO()
    p = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter
    
    y = height - inch # Posición Y actual
    
    def draw_line(text, value, indent=1.2):
        nonlocal y
        p.drawString(inch * indent, y, f"{text}: {value}")
        y -= 0.25 * inch

    def draw_title(text):
        nonlocal y
        y -= 0.35 * inch # Espacio antes del título
        p.setFont("Helvetica-Bold", 12)
        p.drawString(inch, y, text)
        y -= 0.25 * inch
        p.setFont("Helvetica", 11)

    # Título Principal
    p.setFont("Helvetica-Bold", 14)
    p.drawString(inch, y, f"Reporte REM Consolidado - Periodo {datos['periodo']}")
    y -= 0.5 * inch
    p.setFont("Helvetica", 11)
    
    # === 🟢 REM A11: TAMIZAJES MADRE ===
    draw_title("REM A11 - Tamizajes Madre")
    draw_line("Tamizajes VIH Positivos", datos['rem_a11']['tamizajes_vih_positivos'])
    draw_line("Tamizajes VDRL Positivos", datos['rem_a11']['tamizajes_vdrl_positivos'])
    draw_line("VDRL Positivos con Tratamiento", datos['rem_a11']['tamizajes_vdrl_con_tratamiento'])
    draw_line("Tamizajes Hepatitis B Positivos", datos['rem_a11']['tamizajes_hepb_positivos'])

    # === 🟢 REM A21: PARTO Y CLASIFICACIÓN ===
    draw_title("REM A21 - Modelo Atención Parto")
    draw_line("Total Partos Registrados", datos['rem_a21']['total_partos'])
    
    # Desglose dinámico de Tipo de Parto
    p.drawString(inch * 1.2, y, "Desglose por Tipo de Parto:")
    y -= 0.25 * inch
    for desglose in datos['rem_a21']['desglose_partos']:
        draw_line(desglose['tipo_parto__valor'], desglose['total'], indent=1.4)
    
    # Desglose dinámico de Robson
    p.drawString(inch * 1.2, y, "Desglose por Grupo Robson:")
    y -= 0.25 * inch
    for desglose in datos['rem_a21']['desglose_robson']:
        draw_line(f"Grupo {desglose['clasificacion_robson__grupo']}", desglose['total'], indent=1.4)
        
    draw_line("Partos con Acompañamiento", datos['rem_a21']['partos_con_acompanamiento'])
    
    # === 🟢 REM A24: RECIÉN NACIDO Y CALIDAD ===
    draw_title("REM A24 - Atención Recién Nacido")
    draw_line("Total Recién Nacidos", datos['rem_a24']['total_rn'])
    draw_line("RN con Lactancia (60 min)", datos['rem_a24']['rn_con_lm_60min'])
    draw_line("RN con Bajo Peso (<2500g)", datos['rem_a24']['rn_bajo_peso'])
    draw_line("RN con Apgar < 7 (5 min)", datos['rem_a24']['rn_apgar_bajo'])
    draw_line("RN con Reanimación", datos['rem_a24']['rn_con_reanimacion'])
    draw_line("Profilaxis Vitamina K", datos['rem_a24']['rn_con_vitamina_k'])
    draw_line("Profilaxis Oftálmica", datos['rem_a24']['rn_con_prof_oftalmica'])

    p.showPage()
    p.save()
    
    buffer.seek(0)
    response = HttpResponse(
        buffer,
        content_type='application/pdf',
        headers={'Content-Disposition': f'attachment; filename="rem_{datos["periodo"]}.pdf"'},
    )
    return response